from concurrent.futures import ThreadPoolExecutor, as_completed
from pathlib import Path

import openpyxl
from openpyxl.cell.cell import Cell, MergedCell
from tqdm import tqdm

# --- Import Configuration ---
from config import ExcelConfig, TranslatorConfig
from dictionary import NAME_TERM_TRANSLATIONS
from formatting import wrap_text
from Models import PromptReferences, SourceLine, TranslationPrompt, InvalidHeaderException
from text_utils import normalize_cell, safe_str, strip_whitespace
from translator import GeminiTranslationClient
from translator_helper import get_prompt_refrences, parse_translation_response

# --- Sheet utils ---
expected_header = [
    ExcelConfig.TYPE,
    ExcelConfig.ORIGINAL_SPEAKER,
    ExcelConfig.TRANSLATED_SPEAKER,
    ExcelConfig.SOURCE,
    ExcelConfig.TARGET,
]
filled_rows:int = len(expected_header)

def validate_header_row(sheet):
    """
    Checks if the header row is present and contains the expected headers
    """
    sheet_header = [normalize_cell(cell.value) for cell in sheet[1]]
    if sheet_header != expected_header:
        raise InvalidHeaderException(
                f"Header row has incorrect headers: "
                f"Expected headers: {', '.join(expected_header)}."
                f"File headers: {', '.join(sheet_header)}"
        )


# --- API Calling ---

def request_translations_from_api(
    source_lines: list[SourceLine],
    references: PromptReferences,
    translation_client: GeminiTranslationClient,
) -> dict[int, str]:
    """
    Builds the prompt, calls the Gemini API, and parses the response.

    Returns parsed translations keyed by workbook line number.
    """
    batch_prompt = TranslationPrompt(
        references=references,
        lines=source_lines,
        source_lang=TranslatorConfig.SOURCE_LANGUAGE,
        target_lang=TranslatorConfig.TARGET_LANGUAGE,
    )
    api_lines_numbers = {line.line_number for line in source_lines}
    print(f"Sending {len(source_lines)} lines to Gemini...")
    api_translations = translation_client.translate_batch(batch_prompt)

    return parse_translation_response(
        api_translations,
        api_lines_numbers,
    )


# --- Main Processing Logic  ---

def load_workbook(source_file: Path):
    workbook = openpyxl.load_workbook(source_file)
    sheet = workbook.active
    if sheet: 
        return workbook, sheet
    else:
        raise RuntimeError("Workbook is empty. Skipping.")

def process_workbook(
    source_file: Path,
    output_file: Path,
    translation_client: GeminiTranslationClient | None = None,
    replace_single_term: bool = False
) -> bool:
    """
    Processes a single Excel workbook: reads, translates, and saves.

    Returns True if the file was processed and saved successfully.
    Return False if nothing was done.
    """
    file_name = source_file.name
    workbook, sheet = load_workbook(source_file)
    validate_header_row(sheet)  # If this fails the rest doesn't continue
    data_rows = sheet.iter_rows(min_row=2, min_col=1, max_col=filled_rows)  # All rows, excluding header

    # Initialize variables
    source_lines: list[SourceLine] = []  # Formatted lines for translation -- {line_num, speaker, text} format
    dict_translations: dict[int, str] = {}  # For `replace_from_dict` usage
    row_metadata: list[tuple[int, Cell, str]] = []  # Metadata of rows that need translation
    
    # Read and rows and prepare the data
    for line_number, row in enumerate(data_rows, start=1):
        # Read each row
        message_type_cell, origin_speaker_cell, speaker_cell, source_cell, target_cell = row

        # Check if the translation cell is not a MergedCell
        if isinstance(target_cell, MergedCell):
            print(f"WARNING: A merged translation cell was found in row {line_number+1}")
            print("These can not be wrapped nor written properly, skipping row.")
            continue

        # Converts None to empty string and strips leading whitespace
        source_text = safe_str(source_cell.value)
        translation_cell = safe_str(target_cell.value)
        speaker_info = safe_str(speaker_cell.value)
        origin_speaker_info = safe_str(origin_speaker_cell.value)
        message_type = safe_str(message_type_cell.value).lower()
        speaker = speaker_info or origin_speaker_info
    
        # Check if translation is required
        # Assume that if there's source_text and existing translation is empty or starts with "TRANSLATION_ERROR", it needs translation
        needs_translation = source_text != "" and (
                translation_cell == ""  or translation_cell.startswith("TRANSLATION_ERROR")
        )
        if not needs_translation:
            continue

        if replace_single_term:
            normalized_source_text = strip_whitespace(source_text)
            if normalized_source_text in NAME_TERM_TRANSLATIONS:
                dict_translations[line_number] = NAME_TERM_TRANSLATIONS[normalized_source_text]
                row_metadata.append((line_number, target_cell, message_type))
                continue

        source_lines.append(SourceLine(line_number=line_number, speaker=speaker, text=source_text))
    
        # Save metadata of rows that will be translated
        row_metadata.append((line_number, target_cell, message_type))
    
    if not row_metadata:
        print(f"No translations needed for {file_name}")
        if output_file.exists():
            print("\nExisting file was not modified.")
        else:
            print(f"Saved output to: {output_file}")
            workbook.save(output_file)
            return True
        return False
    
    # Call the API and write the translated rows back
    translation_error_count = 0
    translation_client = translation_client or GeminiTranslationClient()

    if source_lines:
        translation_references = get_prompt_refrences(source_lines)
        parsed_api_translations = request_translations_from_api(
            source_lines=source_lines,
            references=translation_references,
            translation_client=translation_client,
        )
    else:
        parsed_api_translations = {}
        
    for line_number, target_cell, message_type in row_metadata:
        if line_number in parsed_api_translations:
            translated_text = parsed_api_translations[line_number]
        elif replace_single_term and line_number in dict_translations:
            translated_text = dict_translations[line_number]
        else:
            translated_text = "TRANSLATION_ERROR: Something happened. You shouldn't be seeing this."
   
        if translated_text.startswith("TRANSLATION_ERROR"):
            translation_error_count += 1
   
        wrapped = wrap_text(translated_text, file_name, message_type)
        target_cell.value = wrapped
        
    workbook.save(output_file)
    if translation_error_count>0:
        # TODO: Expand this to allow choosing from:
        #  a) erroring out(strict mode)
        #  b) retrying with the full context,
        #  c) retrying with only the failed lines
        #  d) Doing nothing (Warning only, current behaviour)
        print(f"WARNING: {file_name} may require a rerun due to {translation_error_count} translation errors.")
    print(f"Saved output to: {output_file}")
    return True


# --- Orchestrator ---

def process_excel_files_in_folder(
    source_folder_path=TranslatorConfig.SOURCE_FOLDER_PATH,
    output_folder_path=TranslatorConfig.OUTPUT_FOLDER_PATH,
    max_parallel_files=TranslatorConfig.MAX_PARALLEL_FILES,
    replace_single_term=TranslatorConfig.REPLACE_SINGLE_TERM,
):
    """
    Finds and processes Excel files (.xlsx) in a given local folder.
    """
    processed_count = 0
    source_folder = Path(source_folder_path)
    output_folder = Path(output_folder_path)

    if not source_folder.is_dir():
        raise NotADirectoryError(f"Source folder not found at {source_folder}")

    output_folder.mkdir(parents=True, exist_ok=True)

    commu_files = sorted(source_folder.glob("*.xlsx"))
    if not commu_files:
        raise FileNotFoundError(f"No Excel files found in {source_folder}.")

    translation_client = GeminiTranslationClient()

    def process_file(source_file_path: Path) -> bool:
        source_file_name = source_file_path.name
        output_file_path = output_folder / source_file_name
        print(f"\n--- Processing file: {source_file_name} ---")
        return process_workbook(source_file_path, output_file_path, translation_client, replace_single_term)

    worker_count = max(1, max_parallel_files)

    with ThreadPoolExecutor(max_workers=worker_count) as executor:
        futures = {
            executor.submit(process_file, source_file_path): source_file_path for source_file_path in commu_files
        }
        for future in tqdm(as_completed(futures), total=len(futures), desc="Processing files", unit="file"):
            source_file_path = futures[future]
            try:
                if future.result():
                    processed_count += 1
            except Exception as e:
                print(f"Error processing {source_file_path.name}: {e}")

    return processed_count


# --- Run the script ---
if __name__ == "__main__":
    print("Starting Gakumas Commu Excel Batch Translator script...")
    total_processed = process_excel_files_in_folder()
    if total_processed > 0:
        print(f"\nScript finished. Processed {total_processed} files.")
